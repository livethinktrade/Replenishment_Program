# STEP 8 FORMAT SALES SHEET FOR KROGER
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
# from openpyxl.styles import colors
# from openpyxl.styles import Font, Color
from openpyxl.styles.fills import PatternFill


class ReportFormat:

    def __init__(self, file_name):

        self.file_name = file_name

        self.wb = openpyxl.load_workbook(self.file_name)

    def color_fill_background(self, sheet_name, min_row=0, max_row=0, min_col=0, max_col=0, color_fill=None):

        """

        :param sheet_name: str
        :param min_row: int
        :param max_row: int
        :param min_col: int
        :param max_col: int
        :param color_fill: str (arg takes in Hex value for the color)
        :return:
        """

        ws = self.wb[f'{sheet_name}']

        # makes the whole sheet white
        pattern_fill_color = PatternFill(patternType='solid', fgColor=f'{color_fill}')

        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                cell.fill = pattern_fill_color

        self.wb.save(self.file_name)

    def white_back_ground(self, sheet_name, min_row=0, max_row=0, min_col=0, max_col=0):

        ws = self.wb[f'{sheet_name}']

        # makes the whole sheet white
        pattern_fill_white = PatternFill(patternType='solid', fgColor='FFFFFF')

        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                cell.fill = pattern_fill_white

        self.wb.save(self.file_name)

    def thin_outline(self, sheet_name, min_row=0, max_row=0, min_col=0, max_col=0):

        ws = self.wb[f'{sheet_name}']

        bd_thin = Side(style='thin', color="000000")

        border = Border(left=bd_thin, top=bd_thin, right=bd_thin, bottom=bd_thin)

        # outlines main sales table
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                cell.border = border

        self.wb.save(self.file_name)

    def bold_font(self, sheet_name, min_row=0, max_row=0, min_col=0, max_col=0):

        ws = self.wb[f'{sheet_name}']

        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                cell.font = Font(bold=True)

        self.wb.save(self.file_name)

    def merge_cells(self, sheet_name, start, end):

        """

         sheet_name: str

         start: str
                this will take the excel cell location (for example M16, A2, etc..)
         end: str
                same type of string taken by the start argument.

        """

        ws = self.wb[f'{sheet_name}']

        ws.merge_cells(f'{start}:{end}')

        self.wb.save(self.file_name)

    def set_value(self, sheet_name, location, value):

        ws = self.wb[f'{sheet_name}']

        ws[f'{location}'].value = value

        self.wb.save(self.file_name)

    def left_align(self, sheet_name, min_row=0, max_row=0, min_col=0, max_col=0):

        ws = self.wb[f'{sheet_name}']

        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')

        self.wb.save(self.file_name)

    def change_column_width(self, sheet_name, column_letter, width):

        """

        :param sheet_name: str
        :param column_letter: (example would be column 'E')
        :param width: integer
        :return:

        """

        ws = self.wb[f'{sheet_name}']

        ws.column_dimensions[f'{column_letter}'].width = width

        self.wb.save(self.file_name)

    def change_font(self, sheet_name, cell_location, font_size=12, font_color='00000000'):

        """
        :param sheet_name:
        :param cell_location: excel cell location (ie. 'A1', 'C5')
        :param font_size: int
        :param font_color: hex code
        """

        ws = self.wb[f'{sheet_name}']

        font_style = Font(size=f"{font_size}", color=font_color)

        ws[f'{cell_location}'].font = font_style

        self.wb.save(self.file_name)

    def center_align(self, sheet_name, min_row=0, max_row=0, min_col=0, max_col=0):

        ws = self.wb[f'{sheet_name}']

        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        self.wb.save(self.file_name)


class SalesReportFormat:

    '''

    class is designed to format the sales table.

    '''

    def __init__(self, filename, replenishment_len, sales_report_len, store_rank_len, store_setting):

        self.filename = filename

        self.store_setting = store_setting

        self.replenishment_len = replenishment_len

        self.sales_report_len = sales_report_len

        self.store_rank_len = store_rank_len

        # internal report variables
        self.initial_orders = self.store_setting.loc['initial_orders','values']
        self.size_build_up = self.store_setting.loc['size_build_up','values']
        self.high_return = self.store_setting.loc['high_return','values']
        self.store_sales_rank = self.store_setting.loc['store_sales_rank','values']
        self.item_approval = self.store_setting.loc['item_approval','values']
        self.store_program = self.store_setting.loc['store_program', 'values']

        # external report variables
        self.ytd_mask_sales_table = self.store_setting.loc['ytd_mask_sales_table','values']
        self.ytd_womask_sales_table = self.store_setting.loc['ytd_womask_sales_table','values']
        self.item_sales_rank = self.store_setting.loc['item_sales_rank','values']
        self.top20 = self.store_setting.loc['top20','values']
        self.in_season_setting = self.store_setting.loc['In_Season', 'values']

        self.wb = openpyxl.load_workbook(self.filename)

        self.bd = Side(style='thick', color="000000")
        self.bd_thin = Side(style='thin', color="000000")

        self.border = Border(left=self.bd_thin, top=self.bd_thin, right=self.bd_thin, bottom=self.bd_thin)

    def replenishment(self, replenishment_len):

        ws = self.wb['Replenishment']

        pattern_fill_gray = PatternFill(patternType='solid', fgColor='D0CECE')
        pattern_fill_green = PatternFill(patternType='solid', fgColor='E2EFDA')
        pattern_fill_darkblue = PatternFill(patternType='solid', fgColor='002060')
        pattern_fill_salmon = PatternFill(patternType='solid', fgColor='FCE4D6')

        bd_thin = Side(style='thin', color="000000")

        # outlining border

        border = Border(left=bd_thin, top=bd_thin, right=bd_thin, bottom=bd_thin)

        for row in ws.iter_rows(min_row=1, max_row=replenishment_len, min_col=1, max_col=3):
            for cell in row:
                cell.border = border

        # setting columns to certain colors
        for row in ws.iter_rows(min_row=1, max_row=replenishment_len, min_col=1, max_col=1):
            for cell in row:
                cell.fill = pattern_fill_gray

        for row in ws.iter_rows(min_row=1, max_row=replenishment_len, min_col=2, max_col=2):
            for cell in row:
                cell.fill = pattern_fill_green

        for row in ws.iter_rows(min_row=1, max_row=replenishment_len, min_col=3, max_col=3):
            for cell in row:
                cell.fill = pattern_fill_salmon

        for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=3):
            for cell in row:
                cell.fill = pattern_fill_darkblue

        # setting cells to bold

        # sets first row to bold and change color to white
        for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=13):
            for cell in row:
                cell.font = Font(bold=True, color='FFFFFF')

        # sets store number to bold
        for row in ws.iter_rows(min_row=2, max_row=replenishment_len, min_col=1, max_col=1):
            for cell in row:
                cell.font = Font(bold=True)

        # center aligns data
        for row in ws.iter_rows(min_row=1, max_row=replenishment_len, min_col=1, max_col=3):
            for cell in row:
                cell.alignment = Alignment(wrap_text=False, horizontal='center', vertical='center')

        ws.column_dimensions['B'].width = 25

        ws['A1'].value = 'Store'
        ws['B1'].value = 'ITEM'
        ws['C1'].value = 'CASE'

        # add in filter

        ws.auto_filter.ref = f"A1:C{replenishment_len+10}"

        self.wb.save(self.filename)

    def sales_report(self, sales_report_len):

        ws = self.wb['Sales Report']

        # makes the whole sheet white
        pattern_fill_white = PatternFill(patternType='solid', fgColor='FFFFFF')

        for row in ws.iter_rows(min_row=0, max_row=500, min_col=0, max_col=52):
            for cell in row:
                cell.fill = pattern_fill_white

        for row in ws.iter_rows(min_row=1, max_row=sales_report_len, min_col=4, max_col=4):
            for cell in row:
                cell.style = 'Percent'

        for row in ws.iter_rows(min_row=1, max_row=sales_report_len, min_col=7, max_col=7):
            for cell in row:
                cell.style = 'Percent'

        for row in ws.iter_rows(min_row=1, max_row=sales_report_len, min_col=10, max_col=10):
            for cell in row:
                cell.style = 'Percent'

        # o7 = ws['O7']
        # o7.style = 'Percent'

        # o12 = ws['O12']
        # o12.style = 'Percent'

        # color set
        pattern_fill_blue = PatternFill(patternType='solid', fgColor='DDEBF7')
        pattern_fill_gray = PatternFill(patternType='solid', fgColor='D0CECE')
        pattern_fill_green = PatternFill(patternType='solid', fgColor='E2EFDA')


        # changes cell colors
        a1 = ws['A1']
        a1.fill = pattern_fill_gray

        for row in ws.iter_rows(min_row=1, max_row=1, min_col=2, max_col=4):
            for cell in row:
                cell.fill = pattern_fill_green

        # change background color to blue for ytd sales (-mask) and store details
        for row in ws.iter_rows(min_row=1, max_row=1, min_col=5, max_col=7):
            for cell in row:
                cell.fill = pattern_fill_blue

        # for row in ws.iter_rows(min_row=5, max_row=5, min_col=13, max_col=15):
        #     for cell in row:
        #         cell.fill = pattern_fill_blue

        # change background color to blue for ytd sales (+mask)
        # for row in ws.iter_rows(min_row=10, max_row=10, min_col=13, max_col=15):
        #     for cell in row:
        #         cell.fill = pattern_fill_blue

        # change background color to blue for item rank
        for row in ws.iter_rows(min_row=1, max_row=1, min_col=13, max_col=15):
            for cell in row:
                cell.fill = pattern_fill_blue

        # changes cells to be bold
        for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=15):
            for cell in row:
                cell.font = Font(bold=True)

            # bold for store ytd column
        # for row in ws.iter_rows(min_row=5, max_row=6, min_col=13, max_col=15):
        #     for cell in row:
        #         cell.font = Font(bold=True)
        #
        #     # Bold ytd + mas
        # for row in ws.iter_rows(min_row=10, max_row=11, min_col=13, max_col=15):
        #     for cell in row:
        #         cell.font = Font(bold=True)

        #     # Bold item sales rank
        # for row in ws.iter_rows(min_row=14, max_row=15, min_col=13, max_col=15):
        #     for cell in row:
        #         cell.font = Font(bold=True)

        # outlines using thin borders
        bd = Side(style='thick', color="000000")
        bd_thin = Side(style='thin', color="000000")

        border = Border(left=bd_thin, top=bd_thin, right=bd_thin, bottom=bd_thin)

        # outlines main sales table
        for row in ws.iter_rows(min_row=1, max_row=sales_report_len, min_col=1, max_col=10):
            for cell in row:
                cell.border = border

            # outlines store sumary detail
        for row in ws.iter_rows(min_row=1, max_row=2, min_col=13, max_col=15):
            for cell in row:
                cell.border = border

        #     # outlines Store YTD Collumns
        # for row in ws.iter_rows(min_row=5, max_row=7, min_col=13, max_col=15):
        #     for cell in row:
        #         cell.border = border
        #
        #     # outlines Store YTD Collumns
        # # for row in ws.iter_rows(min_row=10, max_row=12, min_col=13, max_col=15):
        # #     for cell in row:
        # #         cell.border = border
        #
        # for row in ws.iter_rows(min_row=14, max_row=25, min_col=13, max_col=15):
        #     for cell in row:
        #         cell.border = border

        # outlines using thick borders for all cells on far left side of sales data table

        border = Border(left=bd, top=bd_thin, right=bd_thin, bottom=bd_thin)

        for row in ws.iter_rows(min_row=1, max_row=sales_report_len, min_col=1, max_col=1):
            for cell in row:
                cell.border = border

        # outlines using thick borders for all cells on far right side of sales data table

        border = Border(left=bd_thin, top=bd_thin, right=bd, bottom=bd_thin)

        for row in ws.iter_rows(min_row=1, max_row=sales_report_len, min_col=10, max_col=10):
            for cell in row:
                cell.border = border

        # outlines using thick borders for all cells on far top side of sales data table

        border = Border(left=bd_thin, top=bd, right=bd_thin, bottom=bd_thin)

        for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=10):
            for cell in row:
                cell.border = border

        # outlines using thick borders for all cells on far bottom side of sales data table

        border = Border(left=bd_thin, top=bd_thin, right=bd_thin, bottom=bd)

        for row in ws.iter_rows(min_row=sales_report_len, max_row=sales_report_len, min_col=1, max_col=10):
            for cell in row:
                cell.border = border

        a1 = ws['A1']
        a1.border = Border(left=bd, top=bd, right=bd_thin, bottom=bd_thin)

        j1 = ws['J1']
        j1.border = Border(left=bd_thin, top=bd, right=bd, bottom=bd_thin)

        a_bottom = ws[f'A{sales_report_len}']
        a_bottom.border = Border(left=bd, top=bd_thin, right=bd_thin, bottom=bd)

        j_bottom = ws[f'J{sales_report_len}']
        j_bottom.border = Border(left=bd_thin, top=bd_thin, right=bd, bottom=bd)

        # for row in ws.iter_rows(min_row=1, max_row=sales_report_len, min_col=4, max_col=4):
        #     for cell in row:
        #         cell.style = 'Percent'

        # change width and height of collumns
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['M'].width = 20
        ws.column_dimensions['N'].width = 25
        ws.column_dimensions['O'].width = 20

        ws.row_dimensions[1].height = 60

        # change text to wrap text and center aligns

        for row in ws.iter_rows(min_row=1, max_row=sales_report_len, min_col=1, max_col=20):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=20):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        # for row in ws.iter_rows(min_row=5, max_row=5, min_col=13, max_col=15):
        #     for cell in row:
        #         cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        ws.auto_filter.ref = f"A1:J{self.sales_report_len+10}"

        self.wb.save(self.filename)

    def top20_format(self):

        ws = self.wb['Sales Report']

        pattern_fill_gray = PatternFill(patternType='solid', fgColor='D0CECE')

        ws.row_dimensions[1].height = 30

        for row in ws.iter_rows(min_row=0, max_row=22, min_col=0, max_col=2):
            for cell in row:
                cell.border = self.border

        ws.merge_cells('A1:B1')
        top_left_cell = ws['A1']
        top_left_cell.value = "Top 20 Stores "

        ws.column_dimensions['B'].width = 20

        for row in ws.iter_rows(min_row=1, max_row=22, min_col=1, max_col=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=2):
            for cell in row:
                cell.fill = pattern_fill_gray

        for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=2):
            for cell in row:
                cell.font = Font(bold=True)

        #outlining store summary
        for row in ws.iter_rows(min_row=1, max_row=2, min_col=13, max_col=15):
            for cell in row:
                cell.border = self.border

        #change width of columns for stores supported table
        ws.column_dimensions['M'].width = 20
        ws.column_dimensions['N'].width = 25
        ws.column_dimensions['O'].width = 20

        # change text to wrap text and center aligns

        for row in ws.iter_rows(min_row=1, max_row=30, min_col=1, max_col=20):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=20):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        self.wb.save(self.filename)

    def item_sales_ranked(self):

        ws = self.wb['Sales Report']

        #merge cells
        ws.merge_cells('M14:M15')
        ws.merge_cells('N14:N15')
        ws.merge_cells('O14:O15')
        ws.merge_cells('P14:P15')
        ws.merge_cells('Q14:Q15')
        ws.merge_cells('R14:R15')

        ws['M14'].value = 'Rank'
        ws['N14'].value = 'Item'
        ws['O14'].value = 'Sales ($)'
        ws['P14'].value = 'Sales Per Active Store'
        ws['Q14'].value = 'Active Stores'
        ws['R14'].value = '% of Annual Sales'


        ws['M16'].value = '1'
        ws['M17'].value = '2'
        ws['M18'].value = '3'
        ws['M19'].value = '4'
        ws['M20'].value = '5'
        ws['M21'].value = '6'
        ws['M22'].value = '7'
        ws['M23'].value = '8'
        ws['M24'].value = '9'
        ws['M25'].value = '10'

        # Bold item sales rank
        for row in ws.iter_rows(min_row=14, max_row=15, min_col=13, max_col=18):
            for cell in row:
                cell.font = Font(bold=True)

        bd = Side(style='thick', color="000000")
        bd_thin = Side(style='thin', color="000000")

        border = Border(left=bd_thin, top=bd_thin, right=bd_thin, bottom=bd_thin)

        for row in ws.iter_rows(min_row=14, max_row=25, min_col=13, max_col=18):
            for cell in row:
                cell.border = border

        self.wb.save(self.filename)

    def ytd_mask(self):

        ws = self.wb['Sales Report']

        #merge cells
        ws.merge_cells('M10:M11')
        ws.merge_cells('N10:N11')
        ws.merge_cells('O10:O11')

        ws['M10'].value = 'Division Sales current YTD($) +Mask'
        ws['N10'].value = 'Division Sales previous YTD($) +Mask'
        ws['O10'].value = '% +/- YOY Change'

        o12 = ws['O12']
        o12.style = 'Percent'

        pattern_fill_blue = PatternFill(patternType='solid', fgColor='DDEBF7')

        #change background color
        for row in ws.iter_rows(min_row=10, max_row=10, min_col=13, max_col=15):
            for cell in row:
                cell.fill = pattern_fill_blue

        #change font to bold
        for row in ws.iter_rows(min_row=10, max_row=11, min_col=13, max_col=15):
            for cell in row:
                cell.font = Font(bold=True)

        bd = Side(style='thick', color="000000")
        bd_thin = Side(style='thin', color="000000")

        border = Border(left=bd_thin, top=bd_thin, right=bd_thin, bottom=bd_thin)

        # outlines Store YTD Collumns
        for row in ws.iter_rows(min_row=10, max_row=12, min_col=13, max_col=15):
            for cell in row:
                cell.border = border

        # wrap text allign and center
        for row in ws.iter_rows(min_row=10, max_row=10, min_col=13, max_col=15):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        self.wb.save(self.filename)

    def ytd_womask(self):

        ws = self.wb['Sales Report']

        ws.merge_cells('M5:M6')
        ws.merge_cells('N5:N6')
        ws.merge_cells('O5:O6')

        ws['M5'].value = 'Division Sales current YTD($)'
        ws['N5'].value = 'Division Sales previous YTD($)'
        ws['O5'].value = '% +/- YOY Change'

        o7 = ws['O7']
        o7.style = 'Percent'

        pattern_fill_blue = PatternFill(patternType='solid', fgColor='DDEBF7')

        # change background color
        for row in ws.iter_rows(min_row=5, max_row=5, min_col=13, max_col=15):
            for cell in row:
                cell.fill = pattern_fill_blue

        #change font to bold
        for row in ws.iter_rows(min_row=5, max_row=6, min_col=13, max_col=15):
            for cell in row:
                cell.font = Font(bold=True)

        bd = Side(style='thick', color="000000")
        bd_thin = Side(style='thin', color="000000")

        border = Border(left=bd_thin, top=bd_thin, right=bd_thin, bottom=bd_thin)

        # outlines Store YTD Collumns
        for row in ws.iter_rows(min_row=5, max_row=7, min_col=13, max_col=15):
            for cell in row:
                cell.border = border

        for row in ws.iter_rows(min_row=5, max_row=5, min_col=13, max_col=15):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        self.wb.save(self.filename)

    def no_scans(self, in_season_settings):

        ws = self.wb['No Scan']

        if in_season_settings ==1:

            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 15
            ws['B1'].value = 'Item'

            # outlines main sales table
            for row in ws.iter_rows(min_row=1, max_row=500, min_col=1, max_col=4):
                for cell in row:
                    cell.border = self.border

            ws.auto_filter.ref = f"A1:D10000"

        else:

            ws.column_dimensions['B'].width = 30
            ws['B1'].value = 'Sales $'

            # outlines main sales table
            for row in ws.iter_rows(min_row=1, max_row=1000, min_col=1, max_col=2):
                for cell in row:
                    cell.border = self.border

            ws.auto_filter.ref = f"A1:B10000"




        self.wb.save(self.filename)

    def on_hands(self):

        ws = self.wb['On Hand']

        ws['A1'].value = 'Store'
        ws['B1'].value = 'Item'
        ws['C1'].value = 'Display Size'
        ws['D1'].value = 'Season'
        ws['E1'].value = 'Case Size'
        ws['F1'].value = 'Delivery'
        ws['G1'].value = 'Credit'
        ws['H1'].value = 'Sales'
        ws['I1'].value = 'Bandaid'
        ws['J1'].value = 'On Hand'
        ws['K1'].value = 'Case Qty'

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 15

        # outlines main sales table
        for row in ws.iter_rows(min_row=1, max_row=5000, min_col=1, max_col=11):
            for cell in row:
                cell.border = self.border

        ws.auto_filter.ref = f"A1:K500"

        self.wb.save(self.filename)

    def replenishment_reasons(self):

        ws = self.wb['Replenishment Reasons']

        ws['A1'].value = 'Store'
        ws['B1'].value = 'Display Size'
        ws['C1'].value = 'Item'
        ws['D1'].value = 'Case Qty'
        ws['E1'].value = 'Rejection Reason'

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 100

        ws.auto_filter.ref = f"A1:E10000"

        # outlines main sales table
        for row in ws.iter_rows(min_row=1, max_row=1000, min_col=1, max_col=5):
            for cell in row:
                cell.border = self.border

        self.wb.save(self.filename)

    def initial_orders_tab(self):

        ws = self.wb['Potential Initial Orders']

        ws.column_dimensions['I'].width = 17
        ws.column_dimensions['J'].width = 17
        ws.column_dimensions['N'].width = 18
        ws.column_dimensions['O'].width = 16
        ws.column_dimensions['P'].width = 12
        ws.column_dimensions['Q'].width = 21
        ws.column_dimensions['R'].width = 20

        ws.merge_cells('G1:J1')
        ws['G1'].value = 'Store On Hands'

        ws.merge_cells('M1:R1')
        ws['M1'].value = 'Store On Hand by Display Size'

        # outlines main sales table
        for row in ws.iter_rows(min_row=1, max_row=500, min_col=7, max_col=10):
            for cell in row:
                cell.border = self.border

        for row in ws.iter_rows(min_row=1, max_row=500, min_col=13, max_col=18):
            for cell in row:
                cell.border = self.border

        ws.auto_filter.ref = f"G2:R10000"

        self.wb.save(self.filename)

    def high_returns(self):

        ws = self.wb['High Return']

        ws['A1'].value = 'Store'
        ws['B1'].value = 'Item'
        ws['C1'].value = 'Display Size'
        ws['D1'].value = 'Season'
        ws['E1'].value = 'Case Size'
        ws['F1'].value = 'Delivery'
        ws['G1'].value = 'Credit'
        ws['H1'].value = 'Sales'
        ws['I1'].value = 'On Hand'
        ws['J1'].value = 'Case Qty'
        ws['K1'].value = 'Return %'


        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 15

        for row in ws.iter_rows(min_row=1, max_row=500, min_col=1, max_col=11):
            for cell in row:
                cell.border = self.border

        self.wb.save(self.filename)

    def store_rank(self):

        ws = self.wb['Store Ranks']

        ws['A1'].value = 'Ranking'
        ws['B1'].value = 'Store'
        ws['C1'].value = 'YTD Sales'

        ws.column_dimensions['A'].width = 10

        #numbers the store
        i = 1
        while i <= self.store_rank_len:
            ws[f'A{i+1}'].value = i
            i+=1

        for row in ws.iter_rows(min_row=1, max_row=self.store_rank_len+1, min_col=1, max_col=3):
            for cell in row:
                cell.border = self.border

        ws.auto_filter.ref = f"A1:C10000"

        self.wb.save(self.filename)

    def items_approved(self):

        ws = self.wb['Items Approved']

        ws['D1'].value = 'Display Size'
        ws['E1'].value = 'Item'
        ws['F1'].value = 'Inventory'
        ws['G1'].value = 'Cases Available'
        ws['H1'].value = 'Store Price'

        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 37
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20

        for row in ws.iter_rows(min_row=1, max_row=500, min_col=1, max_col=8):
            for cell in row:
                cell.border = self.border


        ws.auto_filter.ref = f"A1:H300"

        self.wb.save(self.filename)

    def store_programs(self):

        ws = self.wb['Store Program']

        ws['A1'].value = 'Store'
        # ws['C1'].value = 'Long Hanging Top'
        # ws['D1'].value = 'Long Hanging Dress'

        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 9
        ws.column_dimensions['E'].width = 9
        ws.column_dimensions['F'].width = 9
        ws.column_dimensions['G'].width = 9
        ws.column_dimensions['H'].width = 9
        ws.column_dimensions['I'].width = 9
        ws.column_dimensions['J'].width = 9
        ws.column_dimensions['K'].width = 30
        ws.column_dimensions['L'].width = 20


        ws.auto_filter.ref = f"A1:L500"

        for row in ws.iter_rows(min_row=1, max_row=500, min_col=1, max_col=12):
            for cell in row:
                cell.border = self.border

        self.wb.save(self.filename)

    def replenishment_pivot(self):

        pass

    def external_report(self):

        self.replenishment(self.replenishment_len)

        if self.top20 == 1:
            self.top20_format()

        else:
            self.sales_report(self.sales_report_len)

        if self.ytd_mask_sales_table == 1:
            self.ytd_mask()

        if self.ytd_womask_sales_table ==1:
            self.ytd_womask()

        if self.item_sales_rank == 1:

            self.item_sales_ranked()

        self.no_scans(self.in_season_setting)

        self.on_hands()

    def internal_report(self):

        self.replenishment(self.replenishment_len)

        self.on_hands()

        self.replenishment_reasons()

        if self.initial_orders == 1:
            self.initial_orders_tab()

        if self.high_return == 1:
            self.high_returns()

        if self.store_sales_rank == 1:
            self.store_rank()

        if self.item_approval == 1:
            self.items_approved()

        if self.store_program == 1:
            self.store_programs()


class KrogerCorporateFormat(ReportFormat):

    """Formats Kroger Corporate Report"""
    def __init__(self, file_name, corporate_period_table_lengths, sales_table_lengths):
        super().__init__(file_name)

        self.corporate_period_table_lengths = corporate_period_table_lengths
        self.sales_table_lengths = sales_table_lengths
        self.store_names = self.sales_table_lengths.keys()

    def corporate_overview_tab(self):

        """
        Formats the Corporate Overview tab in the excel file

        There are three sections that is being formatted.

        1) Corporate Summary Overview
        2) Corporate Period Overview
        3) Division Period Overview

        """

        ###############################################################################################################
        # first section of the code is formatting the Corporate Summary Overview table

        sheet_name = 'Corporate Overview'

        # makes the whole sheet white
        self.white_back_ground(sheet_name, max_row=300, max_col=40)

        # outlining sales summary table
        self.thin_outline(sheet_name, min_row=5, max_row=17, min_col=4, max_col=9)

        # outlining kroger period summary
        self.thin_outline(sheet_name,
                          min_row=self.corporate_period_table_lengths['Overall Period Summary']['start_row'],

                          max_row=self.corporate_period_table_lengths['Overall Period Summary']['start_row']
                                 + self.corporate_period_table_lengths['Overall Period Summary']['table_length'],
                          min_col=4,
                          max_col=6
                          )

        # Change Column Names for summary table
        self.set_value(sheet_name, 'D5', 'Division')
        self.set_value(sheet_name, 'E5', 'YTD Sales')
        self.set_value(sheet_name, 'F5', 'Current Week Sales')
        self.set_value(sheet_name, 'G5', 'Previous Week Sales')
        self.set_value(sheet_name, 'H5', '# Stores on Programs')
        self.set_value(sheet_name, 'I5', '# Stores on MBO Programs')

        # change width of columns
        self.change_column_width(sheet_name, 'D', 15)
        self.change_column_width(sheet_name, 'F', 15)
        self.change_column_width(sheet_name, 'G', 20)
        self.change_column_width(sheet_name, 'H', 20)
        self.change_column_width(sheet_name, 'I', 20)

        # center store summary tables
        self.center_align(sheet_name, min_row=5, max_row=18, min_col=4, max_col=9)

        summary_start_row = 6

        for store_names in self.store_names:

            # Setting cell values to cleaner store name ie 'kroger_atlanta' ==> 'Atlanta'
            self.set_value(sheet_name, f'D{summary_start_row}', store_names)
            summary_start_row += 1

        self.color_fill_background(sheet_name, min_row=5, max_row=5, min_col=4, max_col=9, color_fill='4f6228')

        # change font color of column names
        for x in ['D', 'E', 'F', 'G', 'H', 'I']:

            self.change_font(sheet_name, f'{x}5', font_color='00FFFFFF')

        ###############################################################################################################
        # Second section is formatting the corporate period overview

        # find the cell location to set the title
        title_row_location = self.corporate_period_table_lengths[f'Overall Period Summary']['start_row'] - 1

        # giving each division period table a title
        self.merge_cells(sheet_name, f'D{title_row_location}', f'F{title_row_location}')
        self.set_value(sheet_name, f'D{title_row_location}', 'Corporate Period Overview')

        self.change_font(sheet_name, f'D{title_row_location}', font_size=14)
        self.bold_font(sheet_name, min_row=title_row_location, max_row=title_row_location, min_col=4, max_col=4)

        # Changing column names
        table_col_names_row_location = self.corporate_period_table_lengths['Overall Period Summary']['start_row']

        self.set_value(sheet_name, f'D{table_col_names_row_location}', 'Kroger Period')
        self.set_value(sheet_name, f'E{table_col_names_row_location}', 'Sales')
        self.set_value(sheet_name, f'F{table_col_names_row_location}', 'Total Units Sold')

        # Changing background color of col names
        self.color_fill_background(sheet_name,
                                   min_row=table_col_names_row_location,
                                   max_row=table_col_names_row_location,
                                   min_col=4,
                                   max_col=6,
                                   color_fill='4f6228')

        # Change Font Color the col names
        for x in ['D', 'E', 'F']:
            self.change_font(sheet_name, f'{x}{table_col_names_row_location}', font_color='00FFFFFF')

        ###############################################################################################################
        # third section is formatting the Division period overview for each division

        for store_names in self.store_names:

            # outlining sales by period table for each division
            self.thin_outline(sheet_name,
                              min_row=self.corporate_period_table_lengths[f'{store_names}']['start_row'],

                              max_row=self.corporate_period_table_lengths[f'{store_names}']['start_row']
                                      + self.corporate_period_table_lengths[f'{store_names}']['table_length'],
                              min_col=4,
                              max_col=6
                              )

            # find the cell location to set the title
            title_row_location = self.corporate_period_table_lengths[f'{store_names}']['start_row']-1

            # giving each division period table a title
            self.merge_cells(sheet_name, f'D{title_row_location}', f'F{title_row_location}')
            self.set_value(sheet_name, f'D{title_row_location}', store_names)

            self.change_font(sheet_name, f'D{title_row_location}', font_size=14)
            self.bold_font(sheet_name, min_row=title_row_location, max_row=title_row_location, min_col=4, max_col=4)
            
            # Changing column names
            table_col_names_row_location = self.corporate_period_table_lengths[f'{store_names}']['start_row']

            self.set_value(sheet_name, f'D{table_col_names_row_location}', 'Kroger Period')
            self.set_value(sheet_name, f'E{table_col_names_row_location}', 'Sales')
            self.set_value(sheet_name, f'F{table_col_names_row_location}', 'Total Units Sold')

            # Changing background color of col names
            self.color_fill_background(sheet_name,
                                       min_row=table_col_names_row_location,
                                       max_row=table_col_names_row_location,
                                       min_col=4,
                                       max_col=6,
                                       color_fill='4f6228')

            # Change Font Color the col names
            for x in ['D', 'E', 'F']:
                self.change_font(sheet_name, f'{x}{table_col_names_row_location}', font_color='00FFFFFF')

        # center align period tables for all division
        self.center_align(sheet_name,
                          min_row=20,
                          max_row=self.corporate_period_table_lengths['max row'],
                          min_col=4,
                          max_col=6)

    def sales_table(self):

        """formats sales table"""

        for store_name in self.store_names:

            self.white_back_ground(store_name, max_row=100, max_col=40)

            self.thin_outline(store_name,
                              max_row=self.sales_table_lengths[f'{store_name}'],
                              max_col=5)

            # change column names
            self.set_value(store_name, 'A1', 'Store')
            self.set_value(store_name, 'B1', 'YTD Sales')
            self.set_value(store_name, 'C1', 'YTD Qty Sold')
            self.set_value(store_name, 'D1', 'Current Week Sales')
            self.set_value(store_name, 'E1', 'Previous Week Sales')

            self.bold_font(store_name, max_row=1, max_col=5)

            self.center_align(store_name,
                              max_row=self.sales_table_lengths[f'{store_name}'],
                              max_col=5)

            # change column width
            self.change_column_width(store_name, 'A', 15)
            self.change_column_width(store_name, 'D', 15)
            self.change_column_width(store_name, 'E', 15)

    def kroger_corporate_format(self):

        """Formats the Entire Kroger Corporate Excel File"""

        self.corporate_overview_tab()

        self.sales_table()


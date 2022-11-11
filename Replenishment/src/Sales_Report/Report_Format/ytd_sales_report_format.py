# STEP 8 FORMAT SALES SHEET FOR KROGER
import openpyxl
from openpyxl.styles import NamedStyle, Font, Border, Side, colors,Color, Alignment
# from openpyxl.styles import colors
# from openpyxl.styles import Font, Color
from openpyxl.styles.fills import PatternFill


# def bold_outline(minrow,maxrow, mincol, maxcol):
#     #bold left side
#     border = Border(left=bd, top=bd_thin, right=bd_thin, bottom=bd_thin)

#     for row in ws.iter_rows(min_row=minrow, max_row=maxrow, min_col=mincol, max_col=maxcol):
#         for cell in row:
#             cell.border = border

#     #bold right side
#     border = Border(left=bd, top=bd_thin, right=bd_thin, bottom=bd_thin)

#     for row in ws.iter_rows(min_row=minrow, max_row=maxrow, min_col=mincol, max_col=maxcol):
#         for cell in row:
#             cell.border = border

#     #bold top

#     border = Border(left=bd_thin, top=bd, right=bd_thin, bottom=bd_thin)

#     for row in ws.iter_rows(min_row=minrow, max_row=maxrow, min_col=mincol, max_col=maxcol):
#         for cell in row:
#             cell.border = border

#     #bold bottom

#     border = Border(left=bd_thin, top=bd_thin, right=bd_thin, bottom=bd)

#     for row in ws.iter_rows(min_row=minrow, max_row=maxrow, min_col=mincol, max_col=maxcol):
#         for cell in row:
#             cell.border = border

#     #bolding all of the corners
#     a1 = ws.cell(minrow,mincol)
#     a1.border = Border(left=bd, top=bd, right=bd_thin, bottom=bd_thin)

#     j1 = ws.cell(minrow,maxcol)
#     j1.border = Border(left=bd_thin, top=bd, right=bd, bottom=bd_thin)

#     a_bottom = ws.cell(maxrow,mincol)
#     a_bottom.border = Border(left=bd, top=bd_thin, right=bd_thin, bottom=bd)

#     j_bottom = ws.cell(maxrow,maxcol)
#     j_bottom.border = Border(left=bd_thin, top=bd_thin, right=bd, bottom=bd)

def ytd_sales_report_format(filename, replenishment_len, sales_report_len):
    wb = openpyxl.load_workbook(filename)
    ws = wb['Sales Report']

    #color set
    pattern_fill_blue = PatternFill(patternType='solid', fgColor= 'DDEBF7')
    pattern_fill_gray = PatternFill(patternType='solid', fgColor= 'D0CECE')
    pattern_fill_green = PatternFill(patternType='solid', fgColor= 'E2EFDA')
    pattern_fill_darkblue = PatternFill(patternType='solid', fgColor= '002060')
    pattern_fill_salmon = PatternFill(patternType='solid', fgColor= 'FCE4D6')
    pattern_fill_white = PatternFill(patternType='solid', fgColor= 'FFFFFF')


    bd = Side(style='thick', color="000000")
    bd_thin = Side(style='thin', color="000000")
    border = Border(left=bd_thin, top=bd_thin, right=bd_thin, bottom=bd_thin)


#Making whole page white

    for row in ws.iter_rows(min_row=0, max_row=500, min_col=0, max_col=52):
        for cell in row:
            cell.fill = pattern_fill_white


#CREATING SALES RANK TABLE

    ws.merge_cells('E10:E11')
    ws.merge_cells('F10:F11')
    ws.merge_cells('G10:G11')


    # ws['E10'].value = 'Rank'
    # ws['F10'].value = 'Item'
    # ws['G10'].value = 'Sales ($)'


    ws['E12'].value = '1'
    ws['E13'].value = '2'
    ws['E14'].value = '3'
    ws['E15'].value = '4'
    ws['E16'].value = '5'
    ws['E17'].value = '6'
    ws['E18'].value = '7'
    ws['E19'].value = '8'
    ws['E20'].value = '9'
    ws['E21'].value = '10'


    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 20

    for row in ws.iter_rows(min_row=10, max_row=21, min_col=5, max_col=7):
        for cell in row:
            cell.border = border

    for row in ws.iter_rows(min_row=10, max_row=21, min_col=5, max_col=7):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')



    top_left_cell = ws['E10']
    top_left_cell.value = "Rank"

    top_left_cell = ws['F10']
    top_left_cell.value = "Item"

    top_left_cell = ws['G10']
    top_left_cell.value = "Sales ($)"

    for row in ws.iter_rows(min_row=10, max_row=10, min_col=5, max_col=7):
        for cell in row:
            cell.font = Font(bold=True)


    # bold_outline(minrow=10, maxrow=20, mincol=4, maxcol=6)

# YTD sales table

    ws.merge_cells('E5:E6')
    ws.merge_cells('F5:F6')
    ws.merge_cells('G5:G6')

    ws['E5'].value = 'Division Sales current YTD($)'
    ws['F5'].value = 'Division Sales previous YTD($)'
    ws['G5'].value = '% +/- YOY Change'

    for row in ws.iter_rows(min_row=5, max_row=7, min_col=5, max_col=7):
        for cell in row:
            cell.border = border

    for row in ws.iter_rows(min_row=5, max_row=5, min_col=5, max_col=7):
        for cell in row:
            cell.fill = pattern_fill_blue

    for row in ws.iter_rows(min_row=5, max_row=7, min_col=5, max_col=7):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    for row in ws.iter_rows(min_row=5, max_row=5, min_col=5, max_col=7):
        for cell in row:
            cell.font = Font(bold=True)

    # bold_outline(minrow=4, maxrow=6, mincol=4, maxcol=6)

# STORES SUPPORTED AND STORE
    
    for row in ws.iter_rows(min_row=0, max_row=2, min_col=5, max_col=7):
        for cell in row:
            cell.border = border

    for row in ws.iter_rows(min_row=1, max_row=2, min_col=5, max_col=7):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True,horizontal='center', vertical='center')

    for row in ws.iter_rows(min_row=1, max_row=1, min_col=5, max_col=7):
        for cell in row:
            cell.fill = pattern_fill_blue

    # bold_outline(minrow=0, maxrow=1, mincol=4, maxcol=6)

    
# TOP 20 STORES


    ws.row_dimensions[1].height = 30


    for row in ws.iter_rows(min_row=0, max_row=22, min_col=0, max_col=2):
        for cell in row:
            cell.border = border

    ws.merge_cells('A1:B1')
    top_left_cell = ws['A1']
    top_left_cell.value = "Top 20 Stores "

    ws.column_dimensions['B'].width = 20

    for row in ws.iter_rows(min_row=1, max_row=22, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=2):
        for cell in row:
            cell.fill = pattern_fill_gray

    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=2):
        for cell in row:
            cell.font = Font(bold=True)

    # bold_outline(minrow=0, maxrow=21, mincol=0, maxcol=1)
    
#####FROM THIS POINT ON CODE IS FORMATTING REPLENISHMENT PAGE #######

    ws = wb['Replenishment']

    #outlining border 

    border = Border(left=bd_thin, top=bd_thin, right=bd_thin, bottom=bd_thin)

    for row in ws.iter_rows(min_row=1, max_row=replenishment_len, min_col=1, max_col=3):
        for cell in row:
            cell.border = border


    # setting columns to certain colors
    for row in ws.iter_rows(min_row=1, max_row=replenishment_len, min_col=1, max_col=1):
        for cell in row:
            cell.fill = pattern_fill_gray

    for row in ws.iter_rows(min_row=1, max_row=replenishment_len, min_col=2, max_col=2):
        for cell in row:
            cell.fill = pattern_fill_green

    for row in ws.iter_rows(min_row=1, max_row=replenishment_len, min_col=3, max_col=3):
        for cell in row:
            cell.fill = pattern_fill_salmon

    for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=3):
        for cell in row:
            cell.fill = pattern_fill_darkblue


    #setting cells to bold

        #sets first row to bold and change color to white
    for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=13):
        for cell in row:
            cell.font = Font(bold=True, color='FFFFFF')

        #sets store number to bold 
    for row in ws.iter_rows(min_row=2, max_row=replenishment_len, min_col=1, max_col=1):
        for cell in row:
            cell.font = Font(bold=True)

    # center aligns data
    for row in ws.iter_rows(min_row=1, max_row=replenishment_len, min_col=1, max_col=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=False, horizontal='center', vertical='center')

    ws.column_dimensions['B'].width = 25

    ws['A1'].value = 'Store'
    ws['B1'].value = 'ITEM'
    ws['C1'].value = 'CASE'



    #####FROM THIS POINT ON CODE IS FORMATTING on hand PAGE #######

    ws = wb['On Hand']

    ws['A1'].value = 'Store'
    ws['B1'].value = 'Item'
    ws['C1'].value = 'Display Size'
    ws['D1'].value = 'Season'
    ws['E1'].value = 'Case Size'
    ws['F1'].value = 'Delivery'
    ws['G1'].value = 'Credit'
    ws['H1'].value = 'Sales'
    ws['I1'].value = 'On Hand'
    ws['J1'].value = 'Case Qty'

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15


    #####FROM THIS POINT ON CODE IS FORMATTING No scans PAGE #######

    ws = wb['No Scan']
    ws.column_dimensions['B'].width = 30

    wb.save(filename)

# STEP 8 FORMAT SALES SHEET FOR KROGER
import openpyxl
from openpyxl.styles import NamedStyle, Font, Border, Side, colors,Color, Alignment
# from openpyxl.styles import colors
# from openpyxl.styles import Font, Color
from openpyxl.styles.fills import PatternFill

class ReportFormat:

    def __init__(self, filename, replenishment_len, sales_report_len, store_rank_len, store_setting):

        self.filename = filename

        self.store_setting = store_setting

        self.replenishment_len = replenishment_len

        self.sales_report_len = sales_report_len

        self.store_rank_len = store_rank_len

        #internal report variables
        self.initial_orders = self.store_setting.loc['initial_orders','values']
        self.size_build_up = self.store_setting.loc['size_build_up','values']
        self.high_return = self.store_setting.loc['high_return','values']
        self.store_sales_rank = self.store_setting.loc['store_sales_rank','values']
        self.item_approval = self.store_setting.loc['item_approval','values']
        self.store_program = self.store_setting.loc['store_program','values']

        # external report variables
        self.ytd_mask_sales_table = self.store_setting.loc['ytd_mask_sales_table','values']
        self.ytd_womask_sales_table = self.store_setting.loc['ytd_womask_sales_table','values']
        self.item_sales_rank = self.store_setting.loc['item_sales_rank','values']
        self.top20 = self.store_setting.loc['top20','values']
        self.in_season_setting = self.store_setting.loc['In_Season', 'values']

        self.wb = openpyxl.load_workbook(self.filename)

        self.bd = Side(style='thick', color="000000")
        self.bd_thin = Side(style='thin', color="000000")

        self.border = Border(left=self.bd_thin, top=self.bd_thin, right=self.bd_thin, bottom=self.bd_thin)

    def replenishment(self, replenishment_len):

        ws = self.wb['Replenishment']

        pattern_fill_gray = PatternFill(patternType='solid', fgColor='D0CECE')
        pattern_fill_green = PatternFill(patternType='solid', fgColor='E2EFDA')
        pattern_fill_darkblue = PatternFill(patternType='solid', fgColor='002060')
        pattern_fill_salmon = PatternFill(patternType='solid', fgColor='FCE4D6')

        bd_thin = Side(style='thin', color="000000")

        # outlining border

        border = Border(left=bd_thin, top=bd_thin, right=bd_thin, bottom=bd_thin)

        for row in ws.iter_rows(min_row=1, max_row=replenishment_len, min_col=1, max_col=3):
            for cell in row:
                cell.border = border

        # setting columns to certain colors
        for row in ws.iter_rows(min_row=1, max_row=replenishment_len, min_col=1, max_col=1):
            for cell in row:
                cell.fill = pattern_fill_gray

        for row in ws.iter_rows(min_row=1, max_row=replenishment_len, min_col=2, max_col=2):
            for cell in row:
                cell.fill = pattern_fill_green

        for row in ws.iter_rows(min_row=1, max_row=replenishment_len, min_col=3, max_col=3):
            for cell in row:
                cell.fill = pattern_fill_salmon

        for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=3):
            for cell in row:
                cell.fill = pattern_fill_darkblue

        # setting cells to bold

        # sets first row to bold and change color to white
        for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=13):
            for cell in row:
                cell.font = Font(bold=True, color='FFFFFF')

        # sets store number to bold
        for row in ws.iter_rows(min_row=2, max_row=replenishment_len, min_col=1, max_col=1):
            for cell in row:
                cell.font = Font(bold=True)

        # center aligns data
        for row in ws.iter_rows(min_row=1, max_row=replenishment_len, min_col=1, max_col=3):
            for cell in row:
                cell.alignment = Alignment(wrap_text=False, horizontal='center', vertical='center')

        ws.column_dimensions['B'].width = 25

        ws['A1'].value = 'Store'
        ws['B1'].value = 'ITEM'
        ws['C1'].value = 'CASE'

        #add in filter

        ws.auto_filter.ref = f"A1:C{replenishment_len+10}"

        self.wb.save(self.filename)

    def sales_report(self, sales_report_len):

        ws = self.wb['Sales Report']


        # makes the whole sheet white
        pattern_fill_white = PatternFill(patternType='solid', fgColor='FFFFFF')

        for row in ws.iter_rows(min_row=0, max_row=500, min_col=0, max_col=52):
            for cell in row:
                cell.fill = pattern_fill_white

        # Merge cells and assigns a value and color

        # wo mask table
        # ws.merge_cells('M5:M6')
        # ws.merge_cells('N5:N6')
        # ws.merge_cells('O5:O6')

        #with mask
        # ws.merge_cells('M10:M11')
        # ws.merge_cells('N10:N11')
        # ws.merge_cells('O10:O11')

        #item sales rank
        # ws.merge_cells('M14:M15')
        # ws.merge_cells('N14:N15')
        # ws.merge_cells('O14:O15')

        #wo mask
        # ws['M5'].value = 'Division Sales current YTD($)'
        # ws['N5'].value = 'Division Sales previous YTD($)'
        # ws['O5'].value = '% +/- YOY Change'

        # ws['M10'].value = 'Division Sales current YTD($) +Mask'
        # ws['N10'].value = 'Division Sales previous YTD($) +Mask'
        # ws['O10'].value = '% +/- YOY Change'

        # ws['M14'].value = 'Rank'
        # ws['N14'].value = 'Item'
        # ws['O14'].value = 'Sales ($)'

        # ws['M16'].value = '1'
        # ws['M17'].value = '2'
        # ws['M18'].value = '3'
        # ws['M19'].value = '4'
        # ws['M20'].value = '5'
        # ws['M21'].value = '6'
        # ws['M22'].value = '7'
        # ws['M23'].value = '8'
        # ws['M24'].value = '9'
        # ws['M25'].value = '10'

        # formating cell format % $

        for row in ws.iter_rows(min_row=1, max_row=sales_report_len, min_col=4, max_col=4):
            for cell in row:
                cell.style = 'Percent'

        for row in ws.iter_rows(min_row=1, max_row=sales_report_len, min_col=7, max_col=7):
            for cell in row:
                cell.style = 'Percent'

        for row in ws.iter_rows(min_row=1, max_row=sales_report_len, min_col=10, max_col=10):
            for cell in row:
                cell.style = 'Percent'

        # o7 = ws['O7']
        # o7.style = 'Percent'

        # o12 = ws['O12']
        # o12.style = 'Percent'

        # color set
        pattern_fill_blue = PatternFill(patternType='solid', fgColor='DDEBF7')
        pattern_fill_gray = PatternFill(patternType='solid', fgColor='D0CECE')
        pattern_fill_green = PatternFill(patternType='solid', fgColor='E2EFDA')


        # changes cell colors
        a1 = ws['A1']
        a1.fill = pattern_fill_gray

        for row in ws.iter_rows(min_row=1, max_row=1, min_col=2, max_col=4):
            for cell in row:
                cell.fill = pattern_fill_green

        # change background color to blue for ytd sales (-mask) and store details
        for row in ws.iter_rows(min_row=1, max_row=1, min_col=5, max_col=7):
            for cell in row:
                cell.fill = pattern_fill_blue

        # for row in ws.iter_rows(min_row=5, max_row=5, min_col=13, max_col=15):
        #     for cell in row:
        #         cell.fill = pattern_fill_blue

        # change background color to blue for ytd sales (+mask)
        # for row in ws.iter_rows(min_row=10, max_row=10, min_col=13, max_col=15):
        #     for cell in row:
        #         cell.fill = pattern_fill_blue

        # change background color to blue for item rank
        for row in ws.iter_rows(min_row=1, max_row=1, min_col=13, max_col=15):
            for cell in row:
                cell.fill = pattern_fill_blue

        # changes cells to be bold
        for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=15):
            for cell in row:
                cell.font = Font(bold=True)

            # bold for store ytd column
        # for row in ws.iter_rows(min_row=5, max_row=6, min_col=13, max_col=15):
        #     for cell in row:
        #         cell.font = Font(bold=True)
        #
        #     # Bold ytd + mas
        # for row in ws.iter_rows(min_row=10, max_row=11, min_col=13, max_col=15):
        #     for cell in row:
        #         cell.font = Font(bold=True)

        #     # Bold item sales rank
        # for row in ws.iter_rows(min_row=14, max_row=15, min_col=13, max_col=15):
        #     for cell in row:
        #         cell.font = Font(bold=True)

        # outlines using thin borders
        bd = Side(style='thick', color="000000")
        bd_thin = Side(style='thin', color="000000")

        border = Border(left=bd_thin, top=bd_thin, right=bd_thin, bottom=bd_thin)

        # outlines main sales table
        for row in ws.iter_rows(min_row=1, max_row=sales_report_len, min_col=1, max_col=10):
            for cell in row:
                cell.border = border

            # outlines store sumary detail
        for row in ws.iter_rows(min_row=1, max_row=2, min_col=13, max_col=15):
            for cell in row:
                cell.border = border

        #     # outlines Store YTD Collumns
        # for row in ws.iter_rows(min_row=5, max_row=7, min_col=13, max_col=15):
        #     for cell in row:
        #         cell.border = border
        #
        #     # outlines Store YTD Collumns
        # # for row in ws.iter_rows(min_row=10, max_row=12, min_col=13, max_col=15):
        # #     for cell in row:
        # #         cell.border = border
        #
        # for row in ws.iter_rows(min_row=14, max_row=25, min_col=13, max_col=15):
        #     for cell in row:
        #         cell.border = border

        # outlines using thick borders for all cells on far left side of sales data table

        border = Border(left=bd, top=bd_thin, right=bd_thin, bottom=bd_thin)

        for row in ws.iter_rows(min_row=1, max_row=sales_report_len, min_col=1, max_col=1):
            for cell in row:
                cell.border = border

        # outlines using thick borders for all cells on far right side of sales data table

        border = Border(left=bd_thin, top=bd_thin, right=bd, bottom=bd_thin)

        for row in ws.iter_rows(min_row=1, max_row=sales_report_len, min_col=10, max_col=10):
            for cell in row:
                cell.border = border

        # outlines using thick borders for all cells on far top side of sales data table

        border = Border(left=bd_thin, top=bd, right=bd_thin, bottom=bd_thin)

        for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=10):
            for cell in row:
                cell.border = border

        # outlines using thick borders for all cells on far bottom side of sales data table

        border = Border(left=bd_thin, top=bd_thin, right=bd_thin, bottom=bd)

        for row in ws.iter_rows(min_row=sales_report_len, max_row=sales_report_len, min_col=1, max_col=10):
            for cell in row:
                cell.border = border

        a1 = ws['A1']
        a1.border = Border(left=bd, top=bd, right=bd_thin, bottom=bd_thin)

        j1 = ws['J1']
        j1.border = Border(left=bd_thin, top=bd, right=bd, bottom=bd_thin)

        a_bottom = ws[f'A{sales_report_len}']
        a_bottom.border = Border(left=bd, top=bd_thin, right=bd_thin, bottom=bd)

        j_bottom = ws[f'J{sales_report_len}']
        j_bottom.border = Border(left=bd_thin, top=bd_thin, right=bd, bottom=bd)

        # for row in ws.iter_rows(min_row=1, max_row=sales_report_len, min_col=4, max_col=4):
        #     for cell in row:
        #         cell.style = 'Percent'

        # change width and height of collumns
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['M'].width = 20
        ws.column_dimensions['N'].width = 25
        ws.column_dimensions['O'].width = 20

        ws.row_dimensions[1].height = 60

        # change text to wrap text and center aligns

        for row in ws.iter_rows(min_row=1, max_row=sales_report_len, min_col=1, max_col=20):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=20):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        # for row in ws.iter_rows(min_row=5, max_row=5, min_col=13, max_col=15):
        #     for cell in row:
        #         cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        ws.auto_filter.ref = f"A1:J{self.sales_report_len+10}"


        self.wb.save(self.filename)

    def top20_format(self):

        ws = self.wb['Sales Report']

        pattern_fill_gray = PatternFill(patternType='solid', fgColor='D0CECE')

        ws.row_dimensions[1].height = 30

        for row in ws.iter_rows(min_row=0, max_row=22, min_col=0, max_col=2):
            for cell in row:
                cell.border = self.border

        ws.merge_cells('A1:B1')
        top_left_cell = ws['A1']
        top_left_cell.value = "Top 20 Stores "

        ws.column_dimensions['B'].width = 20

        for row in ws.iter_rows(min_row=1, max_row=22, min_col=1, max_col=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=2):
            for cell in row:
                cell.fill = pattern_fill_gray

        for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=2):
            for cell in row:
                cell.font = Font(bold=True)

        #outlining store summary
        for row in ws.iter_rows(min_row=1, max_row=2, min_col=13, max_col=15):
            for cell in row:
                cell.border = self.border

        #change width of columns for stores supported table
        ws.column_dimensions['M'].width = 20
        ws.column_dimensions['N'].width = 25
        ws.column_dimensions['O'].width = 20

        # change text to wrap text and center aligns

        for row in ws.iter_rows(min_row=1, max_row=30, min_col=1, max_col=20):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=20):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')


        self.wb.save(self.filename)

    def item_sales_ranked(self):

        ws = self.wb['Sales Report']

        #merge cells
        ws.merge_cells('M14:M15')
        ws.merge_cells('N14:N15')
        ws.merge_cells('O14:O15')
        ws.merge_cells('P14:P15')
        ws.merge_cells('Q14:Q15')
        ws.merge_cells('R14:R15')

        ws['M14'].value = 'Rank'
        ws['N14'].value = 'Item'
        ws['O14'].value = 'Sales ($)'
        ws['P14'].value = 'Sales Per Active Store'
        ws['Q14'].value = 'Active Stores'
        ws['R14'].value = '% of Annual Sales'


        ws['M16'].value = '1'
        ws['M17'].value = '2'
        ws['M18'].value = '3'
        ws['M19'].value = '4'
        ws['M20'].value = '5'
        ws['M21'].value = '6'
        ws['M22'].value = '7'
        ws['M23'].value = '8'
        ws['M24'].value = '9'
        ws['M25'].value = '10'

        # Bold item sales rank
        for row in ws.iter_rows(min_row=14, max_row=15, min_col=13, max_col=18):
            for cell in row:
                cell.font = Font(bold=True)

        bd = Side(style='thick', color="000000")
        bd_thin = Side(style='thin', color="000000")

        border = Border(left=bd_thin, top=bd_thin, right=bd_thin, bottom=bd_thin)


        for row in ws.iter_rows(min_row=14, max_row=25, min_col=13, max_col=18):
            for cell in row:
                cell.border = border



        self.wb.save(self.filename)

    def ytd_mask(self):

        ws = self.wb['Sales Report']

        #merge cells
        ws.merge_cells('M10:M11')
        ws.merge_cells('N10:N11')
        ws.merge_cells('O10:O11')

        ws['M10'].value = 'Division Sales current YTD($) +Mask'
        ws['N10'].value = 'Division Sales previous YTD($) +Mask'
        ws['O10'].value = '% +/- YOY Change'

        o12 = ws['O12']
        o12.style = 'Percent'

        pattern_fill_blue = PatternFill(patternType='solid', fgColor='DDEBF7')

        #change background color
        for row in ws.iter_rows(min_row=10, max_row=10, min_col=13, max_col=15):
            for cell in row:
                cell.fill = pattern_fill_blue

        #change font to bold
        for row in ws.iter_rows(min_row=10, max_row=11, min_col=13, max_col=15):
            for cell in row:
                cell.font = Font(bold=True)

        bd = Side(style='thick', color="000000")
        bd_thin = Side(style='thin', color="000000")

        border = Border(left=bd_thin, top=bd_thin, right=bd_thin, bottom=bd_thin)

        #outlines Store YTD Collumns
        for row in ws.iter_rows(min_row=10, max_row=12, min_col=13, max_col=15):
            for cell in row:
                cell.border = border

        #wrap text allign and center
        for row in ws.iter_rows(min_row=10, max_row=10, min_col=13, max_col=15):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')


        self.wb.save(self.filename)

    def ytd_womask(self):

        ws = self.wb['Sales Report']

        ws.merge_cells('M5:M6')
        ws.merge_cells('N5:N6')
        ws.merge_cells('O5:O6')

        ws['M5'].value = 'Division Sales current YTD($)'
        ws['N5'].value = 'Division Sales previous YTD($)'
        ws['O5'].value = '% +/- YOY Change'

        o7 = ws['O7']
        o7.style = 'Percent'

        pattern_fill_blue = PatternFill(patternType='solid', fgColor='DDEBF7')

        # change background color
        for row in ws.iter_rows(min_row=5, max_row=5, min_col=13, max_col=15):
            for cell in row:
                cell.fill = pattern_fill_blue

        #change font to bold
        for row in ws.iter_rows(min_row=5, max_row=6, min_col=13, max_col=15):
            for cell in row:
                cell.font = Font(bold=True)

        bd = Side(style='thick', color="000000")
        bd_thin = Side(style='thin', color="000000")

        border = Border(left=bd_thin, top=bd_thin, right=bd_thin, bottom=bd_thin)

        # outlines Store YTD Collumns
        for row in ws.iter_rows(min_row=5, max_row=7, min_col=13, max_col=15):
            for cell in row:
                cell.border = border

        for row in ws.iter_rows(min_row=5, max_row=5, min_col=13, max_col=15):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')


        self.wb.save(self.filename)

    def no_scans(self, in_season_settings):

        ws = self.wb['No Scan']

        if in_season_settings ==1:

            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 15
            ws['B1'].value = 'Item'

            # outlines main sales table
            for row in ws.iter_rows(min_row=1, max_row=500, min_col=1, max_col=4):
                for cell in row:
                    cell.border = self.border

            ws.auto_filter.ref = f"A1:D10000"

        else:

            ws.column_dimensions['B'].width = 30
            ws['B1'].value = 'Sales $'

            # outlines main sales table
            for row in ws.iter_rows(min_row=1, max_row=1000, min_col=1, max_col=2):
                for cell in row:
                    cell.border = self.border

            ws.auto_filter.ref = f"A1:B10000"




        self.wb.save(self.filename)

    def on_hands(self):

        ws = self.wb['On Hand']

        ws['A1'].value = 'Store'
        ws['B1'].value = 'Item'
        ws['C1'].value = 'Display Size'
        ws['D1'].value = 'Season'
        ws['E1'].value = 'Case Size'
        ws['F1'].value = 'Delivery'
        ws['G1'].value = 'Credit'
        ws['H1'].value = 'Sales'
        ws['I1'].value = 'On Hand'
        ws['J1'].value = 'Case Qty'

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 15

        # outlines main sales table
        for row in ws.iter_rows(min_row=1, max_row=5000, min_col=1, max_col=10):
            for cell in row:
                cell.border = self.border

        ws.auto_filter.ref = f"A1:J10000"

        self.wb.save(self.filename)

    def replenishment_reasons(self):

        ws = self.wb['Replenishment Reasons']

        ws['A1'].value = 'Store'
        ws['B1'].value = 'Display Size'
        ws['C1'].value = 'Item'
        ws['D1'].value = 'Case Qty'
        ws['E1'].value = 'Rejection Reason'

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 100

        ws.auto_filter.ref = f"A1:E10000"

        # outlines main sales table
        for row in ws.iter_rows(min_row=1, max_row=1000, min_col=1, max_col=5):
            for cell in row:
                cell.border = self.border

        self.wb.save(self.filename)

    def initial_orders_tab(self):

        ws = self.wb['Potential Initial Orders']

        ws.column_dimensions['I'].width = 17
        ws.column_dimensions['J'].width = 17
        ws.column_dimensions['N'].width = 18
        ws.column_dimensions['O'].width = 16
        ws.column_dimensions['P'].width = 12
        ws.column_dimensions['Q'].width = 21
        ws.column_dimensions['R'].width = 20

        ws.merge_cells('G1:J1')
        ws['G1'].value = 'Store On Hands'

        ws.merge_cells('M1:R1')
        ws['M1'].value = 'Store On Hand by Display Size'

        # outlines main sales table
        for row in ws.iter_rows(min_row=1, max_row=500, min_col=7, max_col=10):
            for cell in row:
                cell.border = self.border

        for row in ws.iter_rows(min_row=1, max_row=500, min_col=13, max_col=18):
            for cell in row:
                cell.border = self.border

        ws.auto_filter.ref = f"G2:R10000"

        self.wb.save(self.filename)

    def high_returns(self):

        ws = self.wb['High Return']

        ws['A1'].value = 'Store'
        ws['B1'].value = 'Item'
        ws['C1'].value = 'Display Size'
        ws['D1'].value = 'Season'
        ws['E1'].value = 'Case Size'
        ws['F1'].value = 'Delivery'
        ws['G1'].value = 'Credit'
        ws['H1'].value = 'Sales'
        ws['I1'].value = 'On Hand'
        ws['J1'].value = 'Case Qty'
        ws['K1'].value = 'Return %'


        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 15

        for row in ws.iter_rows(min_row=1, max_row=500, min_col=1, max_col=11):
            for cell in row:
                cell.border = self.border

        self.wb.save(self.filename)

    def store_rank(self):

        ws = self.wb['Store Ranks']

        ws['A1'].value = 'Ranking'
        ws['B1'].value = 'Store'
        ws['C1'].value = 'YTD Sales'

        ws.column_dimensions['A'].width = 10

        #numbers the store
        i = 1
        while i <= self.store_rank_len:
            ws[f'A{i+1}'].value = i
            i+=1

        for row in ws.iter_rows(min_row=1, max_row=self.store_rank_len+1, min_col=1, max_col=3):
            for cell in row:
                cell.border = self.border

        ws.auto_filter.ref = f"A1:C10000"

        self.wb.save(self.filename)

    def items_approved(self):

        ws = self.wb['Items Approved']

        ws['D1'].value = 'Display Size'
        ws['E1'].value = 'Item'
        ws['F1'].value = 'Inventory'
        ws['G1'].value = 'Cases Available'
        ws['H1'].value = 'Store Price'

        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 37
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20

        for row in ws.iter_rows(min_row=1, max_row=500, min_col=1, max_col=8):
            for cell in row:
                cell.border = self.border


        ws.auto_filter.ref = f"A1:H10000"

        self.wb.save(self.filename)

    def store_programs(self):

        ws = self.wb['Store Program']

        ws['A1'].value = 'Store'
        ws['C1'].value = 'Long Hanging Top'
        ws['D1'].value = 'Long Hanging Dress'

        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 20

        ws.auto_filter.ref = f"A1:F10000"

        for row in ws.iter_rows(min_row=1, max_row=500, min_col=1, max_col=6):
            for cell in row:
                cell.border = self.border

        self.wb.save(self.filename)

    def external_report(self):

        self.replenishment(self.replenishment_len)

        if self.top20 == 1:
            self.top20_format()

        else:
            self.sales_report(self.sales_report_len)

        if self.ytd_mask_sales_table == 1:
            self.ytd_mask()

        if self.ytd_womask_sales_table ==1:
            self.ytd_womask()

        if self.item_sales_rank == 1:

            self.item_sales_ranked()

        self.no_scans(self.in_season_setting)

        self.on_hands()

    def internal_report(self):

        self.replenishment_reasons()

        if self.initial_orders == 1:
            self.initial_orders_tab()

        if self.high_return == 1:
            self.high_returns()

        if self.store_sales_rank == 1:
            self.store_rank()

        if self.item_approval ==1:
            self.items_approved()

        if self.store_program ==1:
            self.store_programs()

def weekly_sales_report_format(filename, replenishment_len, sales_report_len):

    wb = openpyxl.load_workbook(filename)
    ws = wb['Sales Report']

    #Making whole page white
    pattern_fill_white = PatternFill(patternType='solid', fgColor= 'FFFFFF')


    for row in ws.iter_rows(min_row=0, max_row=500, min_col=0, max_col=52):
        for cell in row:
            cell.fill = pattern_fill_white

    #Merge cells and assigns a value and color

    ws.merge_cells('M5:M6')
    ws.merge_cells('N5:N6')
    ws.merge_cells('O5:O6')

    ws.merge_cells('M10:M11')
    ws.merge_cells('N10:N11')
    ws.merge_cells('O10:O11')

    ws.merge_cells('M14:M15')
    ws.merge_cells('N14:N15')
    ws.merge_cells('O14:O15')



    ws['M5'].value = 'Division Sales current YTD($)'
    ws['N5'].value = 'Division Sales previous YTD($)'
    ws['O5'].value = '% +/- YOY Change'

    # ws['M10'].value = 'Division Sales current YTD($) +Mask'
    # ws['N10'].value = 'Division Sales previous YTD($) +Mask'
    # ws['O10'].value = '% +/- YOY Change'

    ws['M14'].value = 'Rank'
    ws['N14'].value = 'Item'
    ws['O14'].value = 'Sales ($)'


    ws['M16'].value = '1'
    ws['M17'].value = '2'
    ws['M18'].value = '3'
    ws['M19'].value = '4'
    ws['M20'].value = '5'
    ws['M21'].value = '6'
    ws['M22'].value = '7'
    ws['M23'].value = '8'
    ws['M24'].value = '9'
    ws['M25'].value = '10'


    #formating cell format % $

    for row in ws.iter_rows(min_row=1, max_row=sales_report_len, min_col=4, max_col=4):
        for cell in row:
            cell.style = 'Percent'

    for row in ws.iter_rows(min_row=1, max_row=sales_report_len, min_col=7, max_col=7):
        for cell in row:
            cell.style = 'Percent'

    for row in ws.iter_rows(min_row=1, max_row=sales_report_len, min_col=10, max_col=10):
        for cell in row:
            cell.style = 'Percent'

    o7 = ws['O7']
    o7.style = 'Percent'

    o12 = ws['O12']
    o12.style = 'Percent'



    #color set
    pattern_fill_blue = PatternFill(patternType='solid', fgColor= 'DDEBF7')
    pattern_fill_gray = PatternFill(patternType='solid', fgColor= 'D0CECE')
    pattern_fill_green = PatternFill(patternType='solid', fgColor= 'E2EFDA')
    pattern_fill_darkblue = PatternFill(patternType='solid', fgColor= '002060')
    pattern_fill_salmon = PatternFill(patternType='solid', fgColor= 'FCE4D6')



    #changes cell colors
    a1 = ws['A1']
    a1.fill = pattern_fill_gray

    for row in ws.iter_rows(min_row=1, max_row=1, min_col=2, max_col=4):
        for cell in row:
            cell.fill = pattern_fill_green


    #change background color to blue for ytd sales (-mask) and store details 
    for row in ws.iter_rows(min_row=1, max_row=1, min_col=5, max_col=7):
        for cell in row:
            cell.fill = pattern_fill_blue


    for row in ws.iter_rows(min_row=5, max_row=5, min_col=13, max_col=15):
        for cell in row:
            cell.fill = pattern_fill_blue


    #change background color to blue for ytd sales (+mask)
    # for row in ws.iter_rows(min_row=10, max_row=10, min_col=13, max_col=15):
    #     for cell in row:
    #         cell.fill = pattern_fill_blue

    #change background color to blue for item rank
    for row in ws.iter_rows(min_row=1, max_row=1, min_col=13, max_col=15):
        for cell in row:
            cell.fill = pattern_fill_blue



    # changes cells to be bold
    for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=15):
        for cell in row:
            cell.font = Font(bold=True)

        #bold for store ytd column
    for row in ws.iter_rows(min_row=5, max_row=6, min_col=13, max_col=15):
        for cell in row:
            cell.font = Font(bold=True)

        # Bold ytd + mas 
    for row in ws.iter_rows(min_row=10, max_row=11, min_col=13, max_col=15):
        for cell in row:
            cell.font = Font(bold=True)

        # Bold item sales rank 
    for row in ws.iter_rows(min_row=14, max_row=15, min_col=13, max_col=15):
        for cell in row:
            cell.font = Font(bold=True)


    #outlines using thin borders 
    bd = Side(style='thick', color="000000")
    bd_thin = Side(style='thin', color="000000")

    border = Border(left=bd_thin, top=bd_thin, right=bd_thin, bottom=bd_thin)
        
        #outlines main sales table
    for row in ws.iter_rows(min_row=1, max_row=sales_report_len, min_col=1, max_col=10):
        for cell in row:
            cell.border = border

        #outlines store sumary detail
    for row in ws.iter_rows(min_row=1, max_row=2, min_col=13, max_col=15):
        for cell in row:
            cell.border = border
            
        #outlines Store YTD Collumns
    for row in ws.iter_rows(min_row=5, max_row=7, min_col=13, max_col=15):
        for cell in row:
            cell.border = border

        #outlines Store YTD Collumns
    # for row in ws.iter_rows(min_row=10, max_row=12, min_col=13, max_col=15):
    #     for cell in row:
    #         cell.border = border

    for row in ws.iter_rows(min_row=14, max_row=25, min_col=13, max_col=15):
        for cell in row:
            cell.border = border

    #outlines using thick borders for all cells on far left side of sales data table

    border = Border(left=bd, top=bd_thin, right=bd_thin, bottom=bd_thin)

    for row in ws.iter_rows(min_row=1, max_row=sales_report_len, min_col=1, max_col=1):
        for cell in row:
            cell.border = border



    #outlines using thick borders for all cells on far right side of sales data table

    border = Border(left=bd_thin, top=bd_thin, right=bd, bottom=bd_thin)

    for row in ws.iter_rows(min_row=1, max_row=sales_report_len, min_col=10, max_col=10):
        for cell in row:
            cell.border = border


    #outlines using thick borders for all cells on far top side of sales data table

    border = Border(left=bd_thin, top=bd, right=bd_thin, bottom=bd_thin)

    for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=10):
        for cell in row:
            cell.border = border

    #outlines using thick borders for all cells on far bottom side of sales data table

    border = Border(left=bd_thin, top=bd_thin, right=bd_thin, bottom=bd)

    for row in ws.iter_rows(min_row=sales_report_len, max_row=sales_report_len, min_col=1, max_col=10):
        for cell in row:
            cell.border = border


    a1 = ws['A1']
    a1.border = Border(left=bd, top=bd, right=bd_thin, bottom=bd_thin)

    j1 = ws['J1']
    j1.border = Border(left=bd_thin, top=bd, right=bd, bottom=bd_thin)

    a_bottom = ws[f'A{sales_report_len}']
    a_bottom.border = Border(left=bd, top=bd_thin, right=bd_thin, bottom=bd)

    j_bottom = ws[f'J{sales_report_len}']
    j_bottom.border = Border(left=bd_thin, top=bd_thin, right=bd, bottom=bd)


    # for row in ws.iter_rows(min_row=1, max_row=sales_report_len, min_col=4, max_col=4):
    #     for cell in row:
    #         cell.style = 'Percent'


    #change width and height of collumns
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['M'].width = 20
    ws.column_dimensions['N'].width = 25
    ws.column_dimensions['O'].width = 20

    ws.row_dimensions[1].height = 60

    #change text to wrap text and center aligns 

    for row in ws.iter_rows(min_row=1, max_row=sales_report_len, min_col=1, max_col=20):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')


    for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=20):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    for row in ws.iter_rows(min_row=5, max_row=5, min_col=13, max_col=15):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    # for row in ws.iter_rows(min_row=10, max_row=10, min_col=13, max_col=15):
    #     for cell in row:
    #         cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')


    #####FROM THIS POINT ON CODE IS FORMATTING REPLENISHMENT PAGE #######

    ws = wb['Replenishment']

    #outlining border 

    border = Border(left=bd_thin, top=bd_thin, right=bd_thin, bottom=bd_thin)

    for row in ws.iter_rows(min_row=1, max_row=replenishment_len, min_col=1, max_col=3):
        for cell in row:
            cell.border = border


    # setting columns to certain colors
    for row in ws.iter_rows(min_row=1, max_row=replenishment_len, min_col=1, max_col=1):
        for cell in row:
            cell.fill = pattern_fill_gray

    for row in ws.iter_rows(min_row=1, max_row=replenishment_len, min_col=2, max_col=2):
        for cell in row:
            cell.fill = pattern_fill_green

    for row in ws.iter_rows(min_row=1, max_row=replenishment_len, min_col=3, max_col=3):
        for cell in row:
            cell.fill = pattern_fill_salmon

    for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=3):
        for cell in row:
            cell.fill = pattern_fill_darkblue


    #setting cells to bold

        #sets first row to bold and change color to white
    for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=13):
        for cell in row:
            cell.font = Font(bold=True, color='FFFFFF')

        #sets store number to bold 
    for row in ws.iter_rows(min_row=2, max_row=replenishment_len, min_col=1, max_col=1):
        for cell in row:
            cell.font = Font(bold=True)

    # center aligns data
    for row in ws.iter_rows(min_row=1, max_row=replenishment_len, min_col=1, max_col=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=False, horizontal='center', vertical='center')

    ws.column_dimensions['B'].width = 25

    ws['A1'].value = 'Store'
    ws['B1'].value = 'ITEM'
    ws['C1'].value = 'CASE'


    #####FROM THIS POINT ON CODE IS FORMATTING on hand PAGE #######

    ws = wb['On Hand']

    ws['A1'].value = 'Store'
    ws['B1'].value = 'Item'
    ws['C1'].value = 'Display Size'
    ws['D1'].value = 'Season'
    ws['E1'].value = 'Case Size'
    ws['F1'].value = 'Delivery'
    ws['G1'].value = 'Credit'
    ws['H1'].value = 'Sales'
    ws['I1'].value = 'On Hand'
    ws['J1'].value = 'Case Qty'

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15


    #####FROM THIS POINT ON CODE IS FORMATTING No scans PAGE #######

    ws = wb['No Scan']
    ws.column_dimensions['B'].width = 30

    wb.save(filename)
